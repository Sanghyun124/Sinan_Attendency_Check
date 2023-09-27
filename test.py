import os
import openpyxl

attendence=['O', '0', 'o', '']

path_dir='D:/git/Sinan_Attendency_Check/archive'

file_list=os.listdir(path_dir)
attendencyDict={}

days=len(file_list)

nowWB=openpyxl.load_workbook(path_dir+'/'+'팀셀보고서 9.3.xlsx')
#print(nowWB.sheetnames)
for t in nowWB.sheetnames:
    print(t)
    nowWS=nowWB[t]
    #print(t,nowWS['B6'].value)
    start=6
    while nowWS['B'+str(start)].value:
        name=str(nowWS['C'+str(start)].value)[0:2]+nowWS['B'+str(start)].value
        if name in attendencyDict.keys():
            pass
        else:
            attendencyDict[name]=0
        if nowWS['E'+str(start)].value in attendence:
            attendencyDict[name]+=1
        start+=1

print(attendencyDict)