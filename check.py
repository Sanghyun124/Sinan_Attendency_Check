import os
import openpyxl

attendance=['O', '0', 'o', '']

path_dir='D:/git/Sinan_Attendency_Check/archive'

file_list=os.listdir(path_dir)
attendanceDict={}
print(file_list)
days=len(file_list)

for x in file_list:
    print(x)
    nowWB=openpyxl.load_workbook(path_dir+'/'+x,read_only=True)
    # print(nowWB.sheetnames)
    for t in nowWB.sheetnames:
        nowWS = nowWB[t]
        # print(t,nowWS['B6'].value)
        start = 6
        while nowWS['B' + str(start)].value:
            name = str(nowWS['C' + str(start)].value)[0:2] + nowWS['B' + str(start)].value
            if name in attendanceDict.keys():
                pass
            else:
                attendanceDict[name] = 0
            if str(nowWS['E' + str(start)].value)[0] in attendance:
                attendanceDict[name] += 1
            start += 1
    nowWB.close()
    # print(attendanceDict)


newWB=openpyxl.Workbook()
newWS=newWB.create_sheet("attendance")
cnt=1
for p in attendanceDict.keys():
    newWS['A'+str(cnt)].value=p[0:2]
    newWS['B' + str(cnt)].value = p[2:]
    newWS['C' + str(cnt)].value = attendanceDict[p]
    newWS['D' + str(cnt)].value = days
    newWS['E' + str(cnt)].value = round(attendanceDict[p]/days*100,2)
    cnt+=1
newWB.save('D:/git/Sinan_Attendency_Check/check1.xlsx')